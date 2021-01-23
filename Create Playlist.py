"""
Takes songs from Excel file and searches for them in Spotify, if they are not found they will be added
to a sheet called 'Not Found'.
"""
import json
import requests
from config import spotify_userid, spotify_token, song_file, sheet_name
import openpyxl as xl
from exceptions import ResponseException

class CreatePlaylist:
    def __init__(self):
        self.user_id = spotify_userid
        self.spotify_token = spotify_token
        self.song_file = song_file
        self.sheet_name = sheet_name
        self.all_song_info = {}
        self.song_total = 0
        pass

    # Step 1: Grab artist and song
    def get_songs(self):
        wb = xl.load_workbook(self.song_file)
        sheet = wb[self.sheet_name]
        na_sheet = wb.create_sheet('Not Found')
        max_row_sheet = len([row for row in sheet if not all([cell.value is None for cell in row])]) + 1
        self.song_total = max_row_sheet

        k = 1
        for i in range(2, max_row_sheet):
            song = sheet.cell(row=i, column=1).value
            artist = sheet.cell(row=i, column=2).value
            uri = self.get_spotify_uri(song, artist)

            if uri == "N/a":
                na_sheet.cell(row=k, column=1).value = song
                na_sheet.cell(row=k, column=2).value = artist
                k += 1
                wb.save('Old_Playlist.xlsx')
            else:
                self.all_song_info[i] = {
                    "song_name": song,
                    "artist": artist,
                    "spotify_uri": uri
                }
        pass

    # Step 2: Create a new playlist
    def create_playlist(self):
        request_body = json.dumps({
            "name": "Google Play Likes",
            "description": "All likes from Google Play",
            "public": True
        })

        query = "https://api.spotify.com/v1/users/{}/playlists".format(self.user_id)
        response = requests.post(
            query,
            data=request_body,
            headers={
                "Content-Type": "application/json",
                "Authorization": "Bearer {}".format(spotify_token)
            }
        )
        response_playlist_json = response.json()

        return response_playlist_json["id"]

    # Step 3: Search for the song
    def get_spotify_uri(self, song, artist):
        query = "https://api.spotify.com/v1/search?q=track%3A{}+artist%3A{}&type=track&offset=0&limit=20".format(
            song, artist)

        response = requests.get(query,
                                headers={
                                    "Content-Type": "application/json",
                                    "Authorization": "Bearer {}".format(spotify_token)
                                })
        response_json = response.json()
        songs = response_json["tracks"]["items"]
        #Add exception for IndexError: list index out of range - song not found
        if not songs:
            return "N/a"
        else:
            uri = songs[0]["uri"]
            return uri


    # Step 4: Add to the new Spotify playlist
    def add_song_to_playlist(self):
        #populate our songs dictionary
        self.get_songs()

        #collect all of uri
        uris = [info["spotify_uri"]
                for song, info in self.all_song_info.items()]

        #can only add 100 songs to a Playlist at a time
        songChunks = [uris[x:x + 100] for x in range(0, len(uris), 100)]
        totalSongChunks = len(songChunks)
        #create new playlist
        playlist_id = self.create_playlist()

        #add all songs to playlist
        for i in range(totalSongChunks):
            request_data = json.dumps(songChunks[i])
            query = "https://api.spotify.com/v1/playlists/{}/tracks".format(playlist_id)

            response = requests.post(query,
                                     data=request_data,
                                     headers={
                                         "Content-Type": "application/json",
                                         "Authorization": "Bearer {}".format(spotify_token)
                                     })
            #check for valid response
            if response.status_code not in [200, 201]:
                raise ResponseException(response.status_code)
            response_json = response.json()
        return response_json


if __name__ == "__main__":
    cp = CreatePlaylist()
    cp.add_song_to_playlist()
